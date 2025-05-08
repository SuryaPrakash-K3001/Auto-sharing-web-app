# Updated app.py with matched ride removal
from flask import Flask, request, render_template, redirect, url_for, session, flash
from openpyxl import Workbook, load_workbook
import os
import requests
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Change this to a random secret key

# Excel files
EXCEL_FILE = 'rides.xlsx'
USERS_FILE = 'users.xlsx'
ULTRAMSG_INSTANCE_ID = 'instance114129'
ULTRAMSG_TOKEN = 'ukyr87apjysdlgd7'

# Create Excel files if they don't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Whatsapp', 'Departure', 'Destination', 'Timing', 'UserID'])
    wb.save(EXCEL_FILE)

if not os.path.exists(USERS_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(['Username', 'Password', 'Phone'])
    wb.save(USERS_FILE)

def send_whatsapp_message(to, message):
    url = f"https://api.ultramsg.com/{ULTRAMSG_INSTANCE_ID}/messages/chat"
    payload = {
        "token": ULTRAMSG_TOKEN,
        "to": to,
        "body": message
    }
    try:
        requests.post(url, data=payload)
    except Exception as e:
        print("Failed to send message:", e)

@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        wb = load_workbook(USERS_FILE)
        ws = wb.active
        
        user = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                user = {
                    "username": row[0],
                    "password": row[1],
                    "phone": row[2]
                }
                break
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['username']
            session['phone'] = user['phone']
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password')
    
    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        phone = request.form.get('phone')
        
        if not all([username, password, phone]):
            flash('All fields are required')
            return redirect(url_for('signup'))
        
        wb = load_workbook(USERS_FILE)
        ws = wb.active
        
        # Check if username already exists
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                flash('Username already exists')
                return redirect(url_for('signup'))
        
        # Add new user
        ws.append([username, generate_password_hash(password), phone])
        wb.save(USERS_FILE)
        
        flash('Account created successfully. Please login.')
        return redirect(url_for('login'))
    
    return render_template('signup.html')

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('index.html', username=session['user_id'])

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('phone', None)
    return redirect(url_for('login'))

@app.route('/submit', methods=['POST'])
def submit():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    name = request.form.get('name')
    phone = request.form.get('whatsapp') or session.get('phone')
    departure = request.form.get('departure')
    destination = request.form.get('destination')
    time = request.form.get('timing')

    if not all([name, phone, departure, destination, time]):
        return "All fields are required."

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    matched_row = None
    matched_index = None
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[2] == departure and row[3] == destination and row[4] == time and row[5] != session['user_id']:
            matched_row = {
                "name": row[0],
                "phone": row[1],
                "user_id": row[5]
            }
            matched_index = idx
            break

    if matched_row:
        # Remove the matched row from the worksheet
        ws.delete_rows(matched_index)
        
        # Send notifications to both users
        msg = f"You have a ride partner!\nName: {matched_row['name']}\nPhone: {matched_row['phone']}"
        send_whatsapp_message(phone, msg)
        reverse_msg = f"You have a ride partner!\nName: {name}\nPhone: {phone}"
        send_whatsapp_message(matched_row['phone'], reverse_msg)
        
        wb.save(EXCEL_FILE)
        return "Match found and message sent to both users. The matched ride has been removed from the system."
    else:
        # Add new ride only if no match found
        ws.append([name, phone, departure, destination, time, session['user_id']])
        wb.save(EXCEL_FILE)
        return "Data saved. We will notify you when a match is found."

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)